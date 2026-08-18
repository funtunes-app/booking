#!/usr/bin/env python3
"""Import August 2026 entries from the FunTunes Excel sheet into Supabase."""

import json
import re
import sys
import time
import urllib.request
import urllib.error
import openpyxl

SUPA_URL = "https://gsdthdubpvqhlzwagaye.supabase.co"
SUPA_KEY = "sb_publishable_9xM0H5gZ1I69aFlggTMpcw_Svl9pn0E"

SHEET_NAME = "Funzone - August 2026"
SKIP_NAMES = {"", "-", "NO BUSINESS", "REPEATED"}
BATCH_SIZE = 50


def map_mop(raw):
    """Map spreadsheet MOP values to app-standard values."""
    if not raw or str(raw).strip() in ("-", ""):
        return ""
    s = str(raw).strip()
    if "bank" in s.lower() or "hdfc" in s.lower() or "icici" in s.lower() or "upi" in s.lower():
        return "UPI"
    if s.lower() == "cash":
        return "Cash"
    return s


def parse_amount(val):
    """Convert amount cell to integer. '-', None, empty → 0."""
    if val is None or str(val).strip() in ("-", ""):
        return 0
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return 0


def parse_date(val):
    """Convert DD.MM.YYYY string to YYYY-MM-DD."""
    if not val:
        return None
    s = str(val).strip()
    m = re.match(r"(\d{1,2})\.(\d{1,2})\.(\d{4})", s)
    if m:
        return f"{m.group(3)}-{m.group(2).zfill(2)}-{m.group(1).zfill(2)}"
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return s[:10]
    return None


def parse_phone(val):
    """Convert phone number (may be float) to 10-digit string."""
    if not val:
        return ""
    s = str(val).strip().replace(".0", "").replace(" ", "").replace("+91", "")
    digits = re.sub(r"\D", "", s)
    if len(digits) == 12 and digits.startswith("91"):
        digits = digits[2:]
    if len(digits) == 10:
        return digits
    return ""


def parse_hours(val):
    """Convert hours to string."""
    if not val or str(val).strip() in ("-", ""):
        return "1"
    try:
        h = float(val)
        if h == int(h):
            return str(int(h))
        return str(h)
    except (ValueError, TypeError):
        return "1"


def compute_payment_cols(amount, mop, socks, socks_mop):
    """Split amounts into UPI/Cash columns based on MOP."""
    play_upi = amount if mop == "UPI" else 0
    play_cash = amount if mop == "Cash" else 0
    socks_upi = socks if socks_mop == "UPI" else 0
    socks_cash = socks if socks_mop == "Cash" else 0
    return play_upi, play_cash, socks_upi, socks_cash


def insert_batch(rows):
    """Insert a batch of rows into Supabase via REST API."""
    url = SUPA_URL + "/rest/v1/entries"
    data = json.dumps(rows).encode("utf-8")
    req = urllib.request.Request(url, data=data, method="POST")
    req.add_header("apikey", SUPA_KEY)
    req.add_header("Authorization", f"Bearer {SUPA_KEY}")
    req.add_header("Content-Type", "application/json")
    req.add_header("Prefer", "return=minimal")
    try:
        resp = urllib.request.urlopen(req)
        return {"success": True, "status": resp.status}
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", errors="replace")
        return {"error": f"HTTP {e.code}: {body}"}


def main():
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else None
    if not xlsx_path:
        print("Usage: python3 import_august_2026.py <path_to_xlsx>")
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[SHEET_NAME]

    current_date = None
    entries = []
    skipped = 0

    for row_idx in range(2, ws.max_row + 1):
        sl_no = ws.cell(row_idx, 1).value
        date_raw = ws.cell(row_idx, 2).value
        name = str(ws.cell(row_idx, 3).value or "").strip()
        amount_raw = ws.cell(row_idx, 4).value
        mop_raw = ws.cell(row_idx, 5).value
        socks_raw = ws.cell(row_idx, 6).value
        socks_mop_raw = ws.cell(row_idx, 7).value
        hours_raw = ws.cell(row_idx, 8).value
        time_raw = ws.cell(row_idx, 9).value
        phone_raw = ws.cell(row_idx, 10).value
        dob_raw = ws.cell(row_idx, 11).value

        if date_raw:
            parsed = parse_date(date_raw)
            if parsed:
                current_date = parsed

        if not current_date:
            skipped += 1
            continue

        clean_name = name.replace("`", "").strip()
        if clean_name.upper() in SKIP_NAMES or clean_name.startswith("Extra "):
            skipped += 1
            continue

        if not clean_name:
            skipped += 1
            continue

        amount = parse_amount(amount_raw)
        socks = parse_amount(socks_raw)
        mop = map_mop(mop_raw)
        socks_mop = map_mop(socks_mop_raw)
        hours = parse_hours(hours_raw)
        phone = parse_phone(phone_raw)
        dob = str(dob_raw).strip() if dob_raw else ""
        timing = str(time_raw).strip() if time_raw else ""

        play_upi, play_cash, socks_upi, socks_cash = compute_payment_cols(
            amount, mop, socks, socks_mop
        )

        mop_display = mop if mop else ""
        socks_mop_display = socks_mop if socks_mop else ""

        entry = {
            "date": current_date,
            "customer_name": clean_name,
            "amount": amount,
            "mop": mop_display,
            "socks": socks,
            "socks_mop": socks_mop_display,
            "play_upi": play_upi,
            "play_cash": play_cash,
            "socks_upi": socks_upi,
            "socks_cash": socks_cash,
            "num_kids": 1,
            "hours": hours,
            "time_in": "",
            "time_out": "",
            "timing": timing,
            "phone": phone,
            "dob": dob,
            "entry_type": "funzone",
        }
        entries.append(entry)

    print(f"Parsed {len(entries)} entries, skipped {skipped} rows")

    if not entries:
        print("No entries to import.")
        sys.exit(0)

    total_inserted = 0
    errors = []
    for i in range(0, len(entries), BATCH_SIZE):
        batch = entries[i : i + BATCH_SIZE]
        result = insert_batch(batch)
        if "error" in result:
            errors.append(f"Batch {i//BATCH_SIZE + 1}: {result['error']}")
            print(f"  ERROR batch {i//BATCH_SIZE + 1}: {result['error']}")
        else:
            total_inserted += len(batch)
            print(f"  Inserted batch {i//BATCH_SIZE + 1} ({len(batch)} rows)")
        time.sleep(0.2)

    print(f"\nDone! Inserted {total_inserted}/{len(entries)} entries.")
    if errors:
        print(f"Errors ({len(errors)}):")
        for e in errors:
            print(f"  - {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
