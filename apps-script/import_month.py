#!/usr/bin/env python3
"""
Import any Funzone month from the FunTunes Excel sheet into Supabase.

Usage:
  python3 import_month.py <xlsx_file>                  # list available sheets
  python3 import_month.py <xlsx_file> "August 2026"    # generate SQL for that month
  python3 import_month.py <xlsx_file> "August 2026" --api  # insert via Supabase REST API

The month argument matches against sheet names like "Funzone - August 2026".
You can pass just "August 2026" or the full sheet name.

The default mode generates a .sql file you can paste into the Supabase SQL Editor.
Use --api to insert directly via the REST API (requires network access to Supabase).
"""

import json
import os
import re
import sys
import time
import urllib.request
import urllib.error
import openpyxl

SUPA_URL = "https://gsdthdubpvqhlzwagaye.supabase.co"
SUPA_KEY = "sb_publishable_9xM0H5gZ1I69aFlggTMpcw_Svl9pn0E"

SKIP_NAMES = {"", "-", "NO BUSINESS", "REPEATED"}
BATCH_SIZE = 50


def map_mop(raw):
    if not raw or str(raw).strip() in ("-", ""):
        return ""
    s = str(raw).strip()
    if "bank" in s.lower() or "hdfc" in s.lower() or "icici" in s.lower() or "upi" in s.lower():
        return "UPI"
    if s.lower() == "cash":
        return "Cash"
    return s


def parse_amount(val):
    if val is None or str(val).strip() in ("-", ""):
        return 0
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return 0


def parse_date(val):
    if not val:
        return None
    if hasattr(val, "strftime"):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    m = re.match(r"(\d{1,2})[./](\d{1,2})[./](\d{4})", s)
    if m:
        return f"{m.group(3)}-{m.group(2).zfill(2)}-{m.group(1).zfill(2)}"
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return s[:10]
    return None


def parse_phone(val):
    if not val:
        return ""
    s = str(val).strip().replace(".0", "").replace(" ", "").replace("+91", "")
    digits = re.sub(r"\D", "", s)
    if len(digits) == 12 and digits.startswith("91"):
        digits = digits[2:]
    if len(digits) == 10:
        return digits
    return ""


def parse_hours(val):
    if not val or str(val).strip() in ("-", ""):
        return "1"
    try:
        h = float(val)
        return str(int(h)) if h == int(h) else str(h)
    except (ValueError, TypeError):
        return "1"


def parse_timing(val):
    """Parse time strings like '6.20-7.20', '5.50 PM-6.50 PM', '02:30 PM to 03:30 PM'."""
    if not val:
        return "", "", ""
    s = str(val).strip()
    if not s or s == "-":
        return "", "", ""

    parts = re.split(r'\s*(?:to|-)\s*', s, maxsplit=1)
    if len(parts) != 2:
        return "", "", s

    def to_24h(t):
        t = t.strip()
        m2 = re.match(r'(\d{1,2})[.:](\d{2})\s*(AM|PM)', t, re.I)
        if m2:
            h = int(m2.group(1))
            mi = m2.group(2)
            ap = m2.group(3).upper()
            if ap == "PM" and h != 12:
                h += 12
            if ap == "AM" and h == 12:
                h = 0
            return f"{h:02d}:{mi}"
        m2 = re.match(r'(\d{1,2})[.:](\d{2})', t)
        if m2:
            h = int(m2.group(1))
            mi = m2.group(2)
            return f"{h:02d}:{mi}"
        return ""

    time_in = to_24h(parts[0])
    time_out = to_24h(parts[1])

    def fmt12(t24):
        if not t24:
            return ""
        hh, mm = t24.split(":")
        h = int(hh)
        ap = "PM" if h >= 12 else "AM"
        h = h % 12
        if h == 0:
            h = 12
        return f"{h:02d}:{mm} {ap}"

    timing_display = f"{fmt12(time_in)} to {fmt12(time_out)}" if time_in and time_out else s
    return time_in, time_out, timing_display


def parse_dob(val):
    """Parse DOB — return empty string for '-', None, or empty. Don't default."""
    if not val:
        return ""
    s = str(val).strip()
    if s in ("-", "None", ""):
        return ""
    m = re.match(r"(\d{1,2})[./](\d{1,2})[./](\d{4})", s)
    if m:
        return f"{m.group(3)}-{m.group(2).zfill(2)}-{m.group(1).zfill(2)}"
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return s[:10]
    return s


def sql_escape(s):
    return "'" + str(s).replace("'", "''") + "'"


def parse_sheet(ws):
    current_date = None
    entries = []
    skipped = 0

    for row_idx in range(2, ws.max_row + 1):
        date_raw = ws.cell(row_idx, 2).value
        name = str(ws.cell(row_idx, 3).value or "").strip().replace("`", "")
        amount_raw = ws.cell(row_idx, 4).value
        mop_raw = ws.cell(row_idx, 5).value
        socks_raw = ws.cell(row_idx, 6).value
        socks_mop_raw = ws.cell(row_idx, 7).value
        hours_raw = ws.cell(row_idx, 8).value
        time_raw = ws.cell(row_idx, 9).value
        phone_raw = ws.cell(row_idx, 10).value
        dob_raw = ws.cell(row_idx, 11).value

        if date_raw:
            parsed = parse_date(date_raw)
            if parsed:
                current_date = parsed

        if not current_date:
            skipped += 1
            continue

        if name.upper() in SKIP_NAMES or name.startswith("Extra ") or not name:
            skipped += 1
            continue

        amount = parse_amount(amount_raw)
        socks = parse_amount(socks_raw)
        mop = map_mop(mop_raw)
        socks_mop = map_mop(socks_mop_raw)
        hours = parse_hours(hours_raw)
        phone = parse_phone(phone_raw)
        dob = parse_dob(dob_raw)
        time_in, time_out, timing = parse_timing(time_raw)

        num_kids = 1

        play_upi = amount if mop == "UPI" else 0
        play_cash = amount if mop == "Cash" else 0
        socks_upi = socks if socks_mop == "UPI" else 0
        socks_cash = socks if socks_mop == "Cash" else 0

        entries.append({
            "date": current_date,
            "customer_name": name,
            "amount": amount,
            "mop": mop,
            "socks": socks,
            "socks_mop": socks_mop,
            "play_upi": play_upi,
            "play_cash": play_cash,
            "socks_upi": socks_upi,
            "socks_cash": socks_cash,
            "num_kids": num_kids,
            "hours": hours,
            "time_in": time_in,
            "time_out": time_out,
            "timing": timing,
            "phone": phone,
            "dob": dob,
            "entry_type": "funzone",
        })

    return entries, skipped


def generate_sql(entries, sheet_name, out_path):
    lines = [
        f"-- FunTunes data import: {sheet_name}",
        f"-- Generated from FunTunes_Cust_details.xlsx",
        "-- Run this in the Supabase SQL Editor",
        "",
        "BEGIN;",
        "",
    ]
    for e in entries:
        lines.append(
            f"INSERT INTO entries (date, customer_name, amount, mop, socks, socks_mop, "
            f"play_upi, play_cash, socks_upi, socks_cash, num_kids, hours, "
            f"time_in, time_out, timing, phone, dob, entry_type) VALUES ("
            f"{sql_escape(e['date'])}, {sql_escape(e['customer_name'])}, {e['amount']}, "
            f"{sql_escape(e['mop'])}, {e['socks']}, {sql_escape(e['socks_mop'])}, "
            f"{e['play_upi']}, {e['play_cash']}, {e['socks_upi']}, {e['socks_cash']}, "
            f"1, {sql_escape(e['hours'])}, {sql_escape(e['time_in'])}, {sql_escape(e['time_out'])}, {sql_escape(e['timing'])}, "
            f"{sql_escape(e['phone'])}, {sql_escape(e['dob'])}, 'funzone');"
        )
    lines += ["", "COMMIT;", f"-- Total: {len(entries)} entries"]

    with open(out_path, "w") as f:
        f.write("\n".join(lines))
    print(f"SQL written to: {out_path}")


def insert_via_api(entries):
    total = 0
    errors = []
    for i in range(0, len(entries), BATCH_SIZE):
        batch = entries[i : i + BATCH_SIZE]
        url = SUPA_URL + "/rest/v1/entries"
        data = json.dumps(batch).encode("utf-8")
        req = urllib.request.Request(url, data=data, method="POST")
        req.add_header("apikey", SUPA_KEY)
        req.add_header("Authorization", f"Bearer {SUPA_KEY}")
        req.add_header("Content-Type", "application/json")
        req.add_header("Prefer", "return=minimal")
        try:
            urllib.request.urlopen(req)
            total += len(batch)
            print(f"  Inserted batch {i//BATCH_SIZE + 1} ({len(batch)} rows)")
        except urllib.error.HTTPError as e:
            body = e.read().decode("utf-8", errors="replace")
            errors.append(f"Batch {i//BATCH_SIZE + 1}: HTTP {e.code}: {body}")
            print(f"  ERROR batch {i//BATCH_SIZE + 1}: HTTP {e.code}")
        time.sleep(0.2)

    print(f"\nInserted {total}/{len(entries)} entries.")
    if errors:
        print(f"Errors ({len(errors)}):")
        for e in errors:
            print(f"  - {e}")
        return False
    return True


def find_sheet(wb, query):
    query_lower = query.lower().strip()
    for name in wb.sheetnames:
        if name.lower() == query_lower:
            return name
        if name.lower() == f"funzone - {query_lower}":
            return name
    return None


def list_funzone_sheets(wb):
    return [n for n in wb.sheetnames if n.lower().startswith("funzone")]


def main():
    if len(sys.argv) < 2:
        print("Usage:")
        print("  python3 import_month.py <xlsx_file>                    # list available months")
        print('  python3 import_month.py <xlsx_file> "July 2026"        # generate SQL file')
        print('  python3 import_month.py <xlsx_file> "July 2026" --api  # insert via API')
        sys.exit(1)

    xlsx_path = sys.argv[1]
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    if len(sys.argv) < 3:
        sheets = list_funzone_sheets(wb)
        print(f"Available Funzone sheets ({len(sheets)}):\n")
        for s in sheets:
            ws = wb[s]
            row_count = ws.max_row - 1
            print(f"  {s}  ({row_count} rows)")
        print(f'\nRun: python3 import_month.py "{xlsx_path}" "Month Year"')
        sys.exit(0)

    month_query = sys.argv[2]
    use_api = "--api" in sys.argv

    sheet_name = find_sheet(wb, month_query)
    if not sheet_name:
        print(f"Sheet not found for: {month_query}")
        print("Available sheets:")
        for s in list_funzone_sheets(wb):
            print(f"  {s}")
        sys.exit(1)

    print(f"Processing: {sheet_name}")
    ws = wb[sheet_name]
    entries, skipped = parse_sheet(ws)
    print(f"Parsed {len(entries)} entries, skipped {skipped} rows")

    if not entries:
        print("No entries to import.")
        sys.exit(0)

    dates = sorted(set(e["date"] for e in entries))
    print(f"Date range: {dates[0]} to {dates[-1]}")

    if use_api:
        success = insert_via_api(entries)
        sys.exit(0 if success else 1)
    else:
        slug = month_query.lower().replace(" ", "_")
        out_path = os.path.join(os.path.dirname(__file__), f"import_{slug}.sql")
        generate_sql(entries, sheet_name, out_path)
        print(f"\nPaste the SQL into the Supabase SQL Editor to import.")


if __name__ == "__main__":
    main()
